# coding=utf-8
import os
from openpyxl import load_workbook
import helper
from random import randint


class EXCEL_TO_CSSCI(object):
    """docstring for  EXCEL_TO_CSSCI"""

    def __init__(self, data_path):
        super(EXCEL_TO_CSSCI, self).__init__()
        self.data_path = data_path
        self.excels_path = []
        self.CSSCI_categories = ['【来源篇名】', '【英文篇名】', '【来源作者】',
                                 '【基    金】', '【期    刊】', '【第一机构】',
                                 '【机构名称】', '【第一作者】', '【中图类号】',
                                 '【年代卷期】', '【关 键 词】', '【基金类别】',
                                 '【参考文献】']
        self.start_page = '南京大学中国社会科学研究评价中心\n数字文献处理系统 版本：2.1\n版权所有 (C) 2000 - 2001 CSSCI Corp.\n \n \n'
        self.CSSCI_seperation_line = '-----------------------------------------------------------------------\n\n'
        self.pseudoDates = [',(010):72-78', ',(010):55-61',
                            ',52(010):78-84', ',(050):104-114', ',35(090):3-9']
        self.pseudoCN = ['I210', 'G622.3', 'G613',
                         'G40-057', 'G423.3', 'G23', 'H319.3']

    def get_excel_path(self):
        file_list = os.listdir(self.data_path)
        for file in file_list:
            if file.endswith('.xlsx'):
                self.excels_path.append(file)

    def convert(self):
        self.get_excel_path()
        print("\n \t \t \t \t \t ************* 开始转换Excel到CSSCI格式的文件" +
              " ************* \t \t \t \t \t \n")
        for excel_path in self.excels_path:
            print('现在转换 \"' + excel_path + '\" 文件。')
            self.convert_single_excel(excel_path)

    def convert_single_excel(self, excel_path):
        cssci_name = helper.mk_file_dir(helper.get_file_path(
            self.data_path, 'download_' + excel_path[:-17] + 'excel_to_cssci.txt'))
        cssci_txt = open(cssci_name, 'w+', encoding='utf-8')
        excel = load_workbook(helper.get_file_path(self.data_path, excel_path))
        sheet_name = excel.sheetnames[0]
        sheet = excel[sheet_name]
        cssci_txt.write(self.start_page)
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                info = helper.none_to_empty_string(
                    sheet.cell(row=i, column=j).value)
                if j == 9:
                    line = self.CSSCI_categories[j - 1] + helper.have_another_choice(
                        info, self.pseudoCN[randint(0, 6)]) + '\n'
                elif j == 10:
                    line = self.CSSCI_categories[j - 1] + info + self.pseudoDates[randint(0, 4)] + '\n'
                elif j == 13:
                    line = self.CSSCI_categories[j - 1] + '\n' + info
                else:
                    line = self.CSSCI_categories[j - 1] + info + '\n'
                cssci_txt.write(line)
            cssci_txt.write(self.CSSCI_seperation_line)
        cssci_txt.close()


# test = Excel_TO_CSSCI(r"C:\Users\Administrator\Desktop\citespace\CNKI_CItation_Converter_For_CiteSpace\th-中考-ti-语文-au-　-ab-　-qw-　-cnki_to_text-data"
        # )
#test = Excel_TO_CSSCI(
#    r"C:\Users\Administrator\Desktop\citespace\CNKI_CItation_Converter_For_CiteSpace\th-中考-ti-语文-cnki_to_text-data")
#test.convert()

